import discord
from discord import app_commands
from discord.ext import commands
import os
from dotenv import load_dotenv
from datetime import datetime
import openpyxl
from flask import Flask, jsonify
import threading
import pandas as pd
from flask_cors import CORS

# Charger le token depuis le fichier .env
load_dotenv()
TOKEN = os.getenv('DISCORD_TOKEN')
CHANNEL_ID = os.getenv('CHANNEL_ID')

intents = discord.Intents.default()
bot = commands.Bot(command_prefix="/", intents=intents)

# Dictionnaire pour stocker les temps de début par utilisateur
start_times = {}

# Nom du fichier Excel
EXCEL_FILE = 'charge_log.xlsx'

# Initialisation du fichier Excel s'il n'existe pas
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["UserID", "Date", "Durée (minutes)"])
        wb.save(EXCEL_FILE)

def format_duree(minutes):
    heures = int(minutes // 60)
    mins = int(minutes % 60)
    if heures > 0:
        return f"{heures}h{mins:02d}"
    else:
        return f"{mins} min"

@bot.event
async def on_ready():
    print(f"Connecté en tant que {bot.user}")
    try:
        synced = await bot.tree.sync()
        print(f"Commandes slash synchronisées : {len(synced)}")
    except Exception as e:
        print(e)
    init_excel()

# Commande /start
@bot.tree.command(name="start", description="Démarre le chronomètre de charge")
async def start(interaction: discord.Interaction):
    if CHANNEL_ID and str(interaction.channel_id) != CHANNEL_ID:
        await interaction.response.send_message("❌ Cette commande n'est autorisée que dans le channel dédié.", ephemeral=True)
        return
    user_id = interaction.user.id
    if user_id in start_times:
        await interaction.response.send_message("⏱️ Tu as déjà démarré une session de charge ! Utilise `/stop` pour l'arrêter.", ephemeral=True)
        return
    start_times[user_id] = datetime.now()
    await interaction.response.send_message("🔋 Chronomètre démarré ! Utilise `/stop` pour l'arrêter.", ephemeral=True)

# Commande /stop
@bot.tree.command(name="stop", description="Arrête le chronomètre et enregistre la session dans Excel")
async def stop(interaction: discord.Interaction):
    if CHANNEL_ID and str(interaction.channel_id) != CHANNEL_ID:
        await interaction.response.send_message("❌ Cette commande n'est autorisée que dans le channel dédié.", ephemeral=True)
        return
    user_id = interaction.user.id
    if user_id not in start_times:
        await interaction.response.send_message("❌ Tu n'as pas démarré de session de charge avec `/start`.", ephemeral=True)
        return
    start_time = start_times.pop(user_id)
    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds() / 60  # en minutes
    date_str = start_time.strftime("%Y-%m-%d")
    # Enregistrement dans Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([str(user_id), date_str, round(duration, 2)])
    wb.save(EXCEL_FILE)
    await interaction.response.send_message(f"✅ Session enregistrée : {format_duree(duration)}.", ephemeral=True)

# Commande /stats
@bot.tree.command(name="stats", description="Affiche la moyenne de temps de charge et le nombre de charges par mois")
async def stats(interaction: discord.Interaction):
    if CHANNEL_ID and str(interaction.channel_id) != CHANNEL_ID:
        await interaction.response.send_message("❌ Cette commande n'est autorisée que dans le channel dédié.", ephemeral=True)
        return
    user_id = str(interaction.user.id)
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    sessions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            sessions.append((row[1], row[2]))
    if not sessions:
        await interaction.response.send_message("Aucune session enregistrée.", ephemeral=True)
        return
    # Moyenne
    total = sum([s[1] for s in sessions])
    moyenne = total / len(sessions)
    # Nombre de charges par mois
    from collections import Counter
    mois = [date[:7] for date, _ in sessions]  # 'YYYY-MM'
    compteur = Counter(mois)
    stats_mois = '\n'.join([f"{m} : {c} fois" for m, c in compteur.items()])
    await interaction.response.send_message(f"📊 Moyenne de charge : {format_duree(moyenne)}\nNombre de charges par mois :\n{stats_mois}", ephemeral=True)

# Commande /history
@bot.tree.command(name="history", description="Affiche l'historique complet des charges (tous utilisateurs)")
async def history(interaction: discord.Interaction):
    # Vérifie si la commande est utilisée dans le bon channel
    if CHANNEL_ID and str(interaction.channel_id) != CHANNEL_ID:
        await interaction.response.send_message("❌ Cette commande n'est autorisée que dans le channel dédié.", ephemeral=True)
        return

    # Charge les sessions depuis le fichier Excel
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        sessions = [row for row in ws.iter_rows(min_row=2, values_only=True)]
    except Exception as e:
        await interaction.response.send_message(f"Erreur lors de la lecture du fichier Excel : {e}", ephemeral=True)
        return

    if not sessions:
        await interaction.response.send_message("Aucune session enregistrée.", ephemeral=True)
        return

    # Construit l'historique complet
    msg_lines = []
    for user_id, date, duree in sessions[::-1]:  # Les plus récentes d'abord
        try:
            user = await bot.fetch_user(int(user_id))
            username = user.display_name if hasattr(user, 'display_name') else user.name
        except Exception:
            username = f"ID:{user_id}"
        msg_lines.append(f"**{username}** | {date} : {format_duree(duree)}")
    msg = "\n".join(msg_lines)

    # Si le message est trop long, envoie d'abord un message puis le fichier Excel en pièce jointe
    if len(msg) > 1900:
        await interaction.response.send_message(
            "L'historique est trop long pour être affiché ici. Voici le fichier Excel en pièce jointe.",
            ephemeral=True
        )
        await interaction.followup.send(file=discord.File(EXCEL_FILE), ephemeral=True)
    else:
        await interaction.response.send_message(f"🕓 **Historique complet des charges (tous utilisateurs)** :\n{msg}", ephemeral=True)

# --- FLASK APP POUR AFFICHAGE WEB ---
app = Flask(__name__)
CORS(app)

@app.route('/charges')
def get_charges():
    try:
        df = pd.read_excel(EXCEL_FILE)
        return df.to_json(orient='records', force_ascii=False)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Lancer Flask dans un thread séparé pour ne pas bloquer le bot Discord

def run_flask():
    app.run(port=5000, debug=False, use_reloader=False)

if __name__ == "__main__":
    flask_thread = threading.Thread(target=run_flask)
    flask_thread.daemon = True
    flask_thread.start()
    bot.run(TOKEN) 